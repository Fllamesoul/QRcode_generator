import openpyxl as xl

class Excel ():

    def __init__(self, filepath):
        self.workbook = xl.load_workbook(filepath)
        self.active_sheet = self.workbook.active
        self.rows = self.active_sheet.max_row           #FIXME


    def get_data_cell(self, num_row: int, num_col: int) -> str :
        return str(self.active_sheet.cell(num_row, num_col).value)
